import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from faker import Faker
import random

# Criar um novo arquivo Excel
workbook = openpyxl.Workbook()

# Selecionar a planilha ativa
worksheet = workbook.active

# Definir os títulos das colunas
columns = ['Nome Paciente', 'Data Nascimento Paciente', 'Gênero do Paciente', 'Faixa etária do Paciente',
           'Nome médico', 'Nome Exame', 'Resultado do Exame', 'Convênio']

# Adicionar os títulos das colunas à planilha
for col_num, col_title in enumerate(columns, 1):
    col_letter = get_column_letter(col_num)
    worksheet[f'{col_letter}1'] = col_title
    worksheet[f'{col_letter}1'].font = Font(bold=True)

# Gerar valores aleatórios para cada coluna
fake = Faker('pt_BR')
genders = ['Feminino', 'Masculino']
ages = ['0-10', '11-20', '21-30', '31-40', '41-50', '51-60', '61-70', '71-80', '81+']
exams = ['Ácido úrico', 'Queratina', 'Hemograma', 'Glicemia', 'Lipidograma']
results = ['Padrão normal', 'Abaixo do padrão', 'Acima do padrão', 'Crítico']
insurances = ['Particular', 'Unimed', 'Clinipam', 'Paraná Clínicas']

for row in range(2, 21):  # Gerar 20 registros
    # Nome Paciente
    worksheet[f'A{row}'] = fake.name()

    # Data Nascimento Paciente
    worksheet[f'B{row}'] = fake.date_of_birth(minimum_age=18, maximum_age=90)

    # Gênero do Paciente
    worksheet[f'C{row}'] = random.choice(genders)

    # Faixa etária do Paciente
    worksheet[f'D{row}'] = random.choice(ages)

    # Nome médico
    worksheet[f'E{row}'] = fake.name()

    # Nome Exame
    worksheet[f'F{row}'] = random.choice(exams)

    # Resultado do Exame
    worksheet[f'G{row}'] = random.choice(results)

    # Convênio
    worksheet[f'H{row}'] = random.choice(insurances)

# Configurar largura automática das colunas
for column_cells in worksheet.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

# Salvar o arquivo Excel
workbook.save('planilha_excel.xlsx')
